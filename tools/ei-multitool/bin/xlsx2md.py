
#!/usr/bin/env python3
"""Dump each sheet of an XLSX file into a single Markdown (.md) file.

Usage: python dump-db.py input.xlsx [output_file]

If output_file is omitted, the output file is created next to the input file using the same base name.
"""

import sys
import os
import argparse
from openpyxl import load_workbook


def sheet_to_markdown(sheet) -> str:
	# Collect rows
	rows = list(sheet.iter_rows(values_only=True))
	if not rows:
		return ""  # empty sheet

	# Determine max columns
	max_cols = max((len(r) for r in rows), default=0)
	# Normalize rows
	norm_rows = [tuple((cell if cell is not None else "") for cell in row) + tuple("")*(max_cols - len(row)) for row in rows]

	# Build table: use first row as header
	header = norm_rows[0]
	body = norm_rows[1:]

	def escape(cell):
		s = str(cell)
		s = s.replace("|", "\\|")
		return s

	lines = []
	# Header
	lines.append("| " + " | ".join(escape(c) for c in header) + " |")
	# Separator
	lines.append("| " + " | ".join("---" for _ in header) + " |")
	# Body
	for row in body:
		lines.append("| " + " | ".join(escape(c) for c in row) + " |")

	return "\n".join(lines)


def main(argv):
	parser = argparse.ArgumentParser(
		prog=os.path.basename(argv[0]) if argv else "eidb2md.py",
		description="Dump each sheet of an XLSX file into a single Markdown (.md) file.",
		epilog=(
			"If output_file is omitted, the output file is created next to the input file using the same base name.\n"
			"Use --outdir to place the generated .md file into a different directory."
		),
	)
	parser.add_argument("input", help="Input .xlsx file")
	parser.add_argument("output", nargs="?", help="Output .md file (optional)")
	parser.add_argument("--outdir", "-d", help="Directory to place the output file (optional)")

	if len(argv) < 2:
		parser.print_help()
		return 1

	args = parser.parse_args(argv[1:])

	input_path = args.input
	# Determine base output path
	if args.output:
		output_path = args.output
	else:
		output_path = os.path.splitext(os.path.abspath(input_path))[0] + ".md"
	# If outdir specified, override directory portion of output_path
	if args.outdir:
		base = os.path.basename(output_path)
		output_path = os.path.join(os.path.abspath(args.outdir), base)
	if not os.path.isfile(input_path):
		print(f"Input file not found: {input_path}")
		return 2

	output_dir = os.path.dirname(os.path.abspath(output_path))
	if output_dir:
		os.makedirs(output_dir, exist_ok=True)

	wb = load_workbook(filename=input_path, data_only=True)
	with open(output_path, "w", encoding="utf-8") as f:
		for index, sheet in enumerate(wb.worksheets):
			if index:
				f.write("\n\n")
			f.write(f"# {sheet.title}\n\n")
			md = sheet_to_markdown(sheet)
			if md:
				f.write(md)
			else:
				f.write("_(empty sheet)_")
	print(f"Wrote: {output_path}")

	return 0


if __name__ == "__main__":
	sys.exit(main(sys.argv))
